
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# =====================================================
# FILE PATHS
# =====================================================

file1 = r"D:\Vedant's Folder\NMIMS\Community Service\GRGVS\data\upload\AKARMA LISTED GUJARAT 278.xlsx"
file2 = r"D:\Vedant's Folder\NMIMS\Community Service\GRGVS\data\upload\civil society organization.xlsx"

output_file = r"D:\Vedant's Folder\NMIMS\Community Service\GRGVS\data\master.xlsx"

# =====================================================
# READ FILE 1
# =====================================================

df1 = pd.read_excel(file1, dtype=str, engine="openpyxl")

df1 = df1.rename(columns={
    "SL No.": "Serial No",
    "Company Name": "Company Name",
    "Email ID": "Email ID",
    "Website": "Website",
    "Number": "Number",
    "Registered Address": "Address",
    "City": "City",
    "State": "State",
    "Unnamed: 9": "Person To Contact",
    "Unnamed: 10": "Email Of That Person"
})

# =====================================================
# READ FILE 2
# =====================================================

df2 = pd.read_excel(file2, dtype=str, engine="openpyxl")

df2 = df2.rename(columns={
    "No.": "Serial No",
    "Name of the NGO": "Company Name",
    "E-mail": "Email ID",
    "Contact No.": "Number",
    "Address": "Address",
    "Contact Person": "Person To Contact"
})

# =====================================================
# STANDARD COLUMNS
# =====================================================

required_columns = [
    "Serial No",
    "Company Name",
    "Email ID",
    "Website",
    "Number",
    "Address",
    "City",
    "State",
    "Person To Contact",
    "Email Of That Person"
]

# =====================================================
# CREATE MISSING COLUMNS
# =====================================================

for col in required_columns:

    if col not in df1.columns:
        df1[col] = ""

    if col not in df2.columns:
        df2[col] = ""

# =====================================================
# FORCE SAME STRUCTURE
# =====================================================

df1 = df1[required_columns]
df2 = df2[required_columns]

# =====================================================
# MERGE FILES
# =====================================================

master_df = pd.concat(
    [df1, df2],
    ignore_index=True
)

# =====================================================
# CLEAN DATA
# =====================================================

for col in master_df.columns:

    master_df[col] = (
        master_df[col]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.replace(r"\s+", " ", regex=True)
    )

# =====================================================
# REMOVE EMPTY ROWS
# =====================================================

master_df = master_df[
    ~(
        master_df["Company Name"].eq("")
        &
        master_df["Address"].eq("")
    )
]

# =====================================================
# CREATE DUPLICATE KEY
# =====================================================

master_df["duplicate_key"] = (
    master_df["Company Name"]
    .str.lower()
    .str.strip()
)

# =====================================================
# REMOVE DUPLICATES
# =====================================================

before = len(master_df)

master_df = master_df.drop_duplicates(
    subset=["duplicate_key"],
    keep="first"
)

after = len(master_df)

removed = before - after

master_df = master_df.drop(
    columns=["duplicate_key"]
)

# =====================================================
# SORT
# =====================================================

master_df = master_df.sort_values(
    by="Company Name",
    ascending=True
)

# =====================================================
# RESET SERIAL NUMBER
# =====================================================

master_df = master_df.reset_index(
    drop=True
)

master_df["Serial No"] = range(
    1,
    len(master_df) + 1
)

# =====================================================
# SAVE EXCEL
# =====================================================

master_df.to_excel(
    output_file,
    index=False
)

# =====================================================
# FORMAT EXCEL
# =====================================================

wb = load_workbook(output_file)
ws = wb.active

# Freeze header
ws.freeze_panes = "A2"

# Header formatting
for cell in ws[1]:
    cell.font = Font(bold=True)

# Center Serial No
for cell in ws["A"]:
    cell.alignment = Alignment(horizontal="center")

# Auto width
for column in ws.columns:

    max_length = 0

    column_letter = column[0].column_letter

    for cell in column:

        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass

    adjusted_width = min(max_length + 5, 60)

    ws.column_dimensions[
        column_letter
    ].width = adjusted_width

wb.save(output_file)

# =====================================================
# SUMMARY
# =====================================================

print("\n====================================")
print("DATABASE CREATED SUCCESSFULLY")
print("====================================")
print(f"Total Records : {len(master_df)}")
print(f"Duplicates Removed : {removed}")
print(f"Saved To : {output_file}")
print("====================================")

